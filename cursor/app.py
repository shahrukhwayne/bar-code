import os
import csv
import io
import zipfile
import base64
import time
import tempfile

from flask import Flask, render_template, request, send_file , after_this_request
from barcode import get_barcode_class
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from concurrent.futures import ThreadPoolExecutor


try:
    SKU_FONT = ImageFont.truetype("font/DejaVuSans-Bold.ttf", 30)
    TITLE_FONT = ImageFont.truetype("font/DejaVuSans-Bold.ttf", 20)
except:
    SKU_FONT = ImageFont.load_default()
    TITLE_FONT = ImageFont.load_default()


def create_app():

    app = Flask(__name__)
    app.secret_key = os.urandom(24)

    os.makedirs("static/barcodes", exist_ok=True)
    progress_status = {
    "current": 0,
    "total": 0
}


# ---------------- READ CSV ---------------- #

    def read_csv(file_bytes):

        text = file_bytes.decode("utf-8-sig", errors="replace")

        stream = io.StringIO(text)

        reader = csv.reader(stream)

        next(reader, None)

        rows = []

        for row in reader:

            if not row or len(row) < 2:
                continue

            sku = row[0].strip()
            title = row[1].strip() if row[1] else ""

            if not sku or sku.lower() == "none":
                continue

            rows.append((sku, title))

        return rows


# ---------------- READ XLSX ---------------- #

    def read_xlsx(file_bytes):

        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)

        ws = wb.active

        rows = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):

            # skip header
            if i == 0:
                continue

            if not row or len(row) < 2:
                continue

            sku = row[0]

            if not sku:
                continue

            sku = str(sku).strip()

            if not sku or sku.lower() == "none":
                continue

            title = "" if row[1] is None else str(row[1]).strip()

            rows.append((sku, title))

        return rows


# ---------------- TEXT WRAP ---------------- #

    def split_text_into_lines(text, max_chars=45):

        words = text.split(" ")

        lines = []

        current = ""

        for word in words:

            if len(current) + len(word) + 1 > max_chars:

                lines.append(current)

                current = word

            else:

                if current:
                    current += " "

                current += word

        if current:
            lines.append(current)

        return lines


# ---------------- BARCODE IMAGE ---------------- #

    def generate_barcode_image(sku, title):

        barcode_class = get_barcode_class("code128")

        options = {
            "module_width": 0.9,
            "module_height": 30,
            "font_size": 0,
            "quiet_zone": 8,
            "dpi": 200
        }

        barcode = barcode_class(sku, writer=ImageWriter())

        buffer = io.BytesIO()

        barcode.write(buffer, options)

        buffer.seek(0)

        barcode_img = Image.open(buffer).convert("RGB")

        padding = 40

        lines = split_text_into_lines(title)

        line_height = TITLE_FONT.getbbox("A")[3]

        title_height = (line_height + 10) * len(lines)

        new_height = barcode_img.height + padding + 60 + title_height + 40

        new_img = Image.new("RGB", (barcode_img.width, new_height), (255, 255, 255))

        new_img.paste(barcode_img, (0, padding))

        draw = ImageDraw.Draw(new_img)

        sku_bbox = SKU_FONT.getbbox(sku)

        sku_width = sku_bbox[2] - sku_bbox[0]

        sku_x = (new_img.width - sku_width) / 2

        sku_y = barcode_img.height + padding + 5

        draw.text((sku_x, sku_y), sku, font=SKU_FONT, fill="black")

        text_y = sku_y + 55

        for line in lines:

            bbox = TITLE_FONT.getbbox(line)

            text_width = bbox[2] - bbox[0]

            text_x = (new_img.width - text_width) / 2

            draw.text((text_x, text_y), line, font=TITLE_FONT, fill="black")

            text_y += (bbox[3] - bbox[1]) + 10

        final = io.BytesIO()

        new_img.save(final, format="PNG")

        final.seek(0)

        return final


# ---------------- PARALLEL WORKER ---------------- #

    def generate_pdf_barcode(data):

        sku, title, index = data

        img_buffer = generate_barcode_image(sku, title)

        pdf_buffer = io.BytesIO()

        Image.open(img_buffer).save(pdf_buffer, "PDF", resolution=150.0)

        pdf_buffer.seek(0)

        safe_name = sku.replace(" ", "_")

        return f"{safe_name}_{index}.pdf", pdf_buffer.read()


# ---------------- HOME ---------------- #

    @app.route("/")
    def index():

        return render_template("index.html", barcode_items=None)


# ---------------- GENERATE ---------------- #

    @app.route("/generate-barcodes", methods=["POST"])
    def generate_barcodes():

        uploaded = request.files.get("file")

        if not uploaded:
            return render_template("index.html", barcode_error="Upload file required")

        file_bytes = uploaded.read()

        if uploaded.filename.endswith(".csv"):
            rows = read_csv(file_bytes)

        elif uploaded.filename.endswith(".xlsx"):
            rows = read_xlsx(file_bytes)

        else:
            return render_template("index.html", barcode_error="Only CSV/XLSX allowed")

        print("UPLOADED ROWS:", len(rows))

        app.config["BARCODE_ROWS"] = rows

        results = []

        if rows:

            sku, title = rows[0]

            img_buffer = generate_barcode_image(sku, title)

            results.append({
                "url": "data:image/png;base64," +
                base64.b64encode(img_buffer.getvalue()).decode(),
                "value": sku,
                "title": title
            })

        return render_template("index.html", barcode_items=results)


# ---------------- DOWNLOAD ---------------- #

    @app.route("/progress")
    def progress():
        return {
            "current": progress_status["current"],
            "total": progress_status["total"]
        }

    @app.route("/download-barcodes")
    def download_barcodes():

        start_time = time.time()

        rows = app.config.get("BARCODE_ROWS")

        if not rows:
            return render_template("index.html", barcode_error="Generate barcodes first")

        progress_status["current"] = 0
        progress_status["total"] = len(rows)

        print("TOTAL ROWS:", len(rows))

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        zip_path = temp_file.name

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_STORED) as zip_file:

            index_data = [(sku, title, i) for i, (sku, title) in enumerate(rows)]

            with ThreadPoolExecutor(max_workers=6) as executor:

                results = executor.map(generate_pdf_barcode, index_data)

                count = 0

                for filename, pdf_data in results:

                    zip_file.writestr(filename, pdf_data)

                    count += 1
                    progress_status["current"] = count

                    if count % 100 == 0:
                        print("Generated:", count)

        end_time = time.time()

        print("TOTAL TIME:", round(end_time - start_time, 2), "seconds")

        @after_this_request
        def cleanup(response):
            try:
                os.remove(zip_path)
            except Exception:
                pass
            return response

        return send_file(
            zip_path,
            mimetype="application/zip",
            as_attachment=True,
            download_name="barcodes.zip"
        )

    return app


app = create_app()


if __name__ == "__main__":

    app.run(debug=True, use_reloader=False)